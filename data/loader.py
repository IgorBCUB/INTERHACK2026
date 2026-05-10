"""Load and normalise data from the Hackathon Excel files."""
from __future__ import annotations
import os
import math
from typing import Dict, List, Tuple
from datetime import datetime, time as dtime

import openpyxl

from vrp.models import Order, Vehicle
import config

# Warehouse coordinates (Mollet del Vallès)
DEPOT_LAT = config.DEPOT["lat"]
DEPOT_LON = config.DEPOT["lon"]


# ── Helpers ────────────────────────────────────────────────────────────────────

def _time_to_seconds(t) -> int:
    """Convert datetime.time or None to seconds from midnight."""
    if t is None:
        return 0
    if isinstance(t, dtime):
        return t.hour * 3600 + t.minute * 60 + t.second
    return 0


def _safe_float(v, default=0.0) -> float:
    try:
        return float(v) if v is not None else default
    except (ValueError, TypeError):
        return default


# ── Time-window loader ─────────────────────────────────────────────────────────

def load_time_windows(path: str) -> Dict[str, Dict[int, Tuple[int, int]]]:
    """
    Returns  {client_id: {day_of_week(1-7): (open_sec, close_sec)}}
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    windows: Dict[str, Dict[int, Tuple[int, int]]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        client = str(row[0]) if row[0] else None
        day = row[4]
        open_t = _time_to_seconds(row[10])
        close_t = _time_to_seconds(row[11])
        closed = str(row[12]).strip().upper() if row[12] else ""
        if not client or day is None:
            continue
        if closed in ("X", "SI", "S", "YES", "Y", "1"):
            continue
        if close_t == 0:
            close_t = 86399
        windows.setdefault(client, {})[int(day)] = (open_t, close_t)
    wb.close()
    return windows


# ── Material weight/volume loader ──────────────────────────────────────────────

def load_material_volumes(path: str) -> Dict[str, float]:
    """Returns {material_code: volume_litres_per_box}"""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    vols: Dict[str, float] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        mat = str(row[0]).strip() if row[0] else None
        uma = str(row[2]).strip().upper() if row[2] else ""
        volume = _safe_float(row[13])
        if mat and uma == "CAJ" and volume > 0:
            vols[mat] = volume
    wb.close()
    return vols


# ── Pallet density loader (cajas por palet, columna Contador con UMA=PAL) ──────

def load_pallet_density(path: str = None) -> Dict[str, int]:
    """
    Lee ZM040.XLSX filtrando UMA=PAL y devuelve {material_code: contador}.
    Ej: density['0CF0054'] = 60 → una caja ocupa 1/60 de palet.
    """
    if path is None:
        path = config.ZM040_XLSX
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    density: Dict[str, int] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        mat = str(row[0]).strip() if row[0] else None
        uma = str(row[2]).strip().upper() if row[2] else ""
        contador = row[3]
        if mat and uma == "PAL" and contador and int(contador) > 0:
            density[mat] = int(contador)
    wb.close()
    return density


# ── Fleet type assignment ──────────────────────────────────────────────────────

def assign_vehicle_types(driver_ids: List[str]) -> Dict[str, str]:
    """
    Asigna a cada conductor un vehicle_type según FLEET_COMPOSITION.
    Determinístico por orden alfabético de driver_id.
    Si hay más conductores que vehículos en la flota, los sobrantes son truck_6 (default).
    """
    composition = config.FLEET_COMPOSITION
    n_truck_8 = composition.get("truck_8", 0)
    n_truck_6 = composition.get("truck_6", 0)
    n_van_3   = composition.get("van_3",   0)

    sorted_ids = sorted(driver_ids)
    assignment: Dict[str, str] = {}
    i = 0
    for did in sorted_ids[:n_truck_8]:
        assignment[did] = "truck_8"
    i = n_truck_8
    for did in sorted_ids[i:i + n_truck_6]:
        assignment[did] = "truck_6"
    i += n_truck_6
    for did in sorted_ids[i:i + n_van_3]:
        assignment[did] = "van_3"
    i += n_van_3
    # Sobrantes
    for did in sorted_ids[i:]:
        assignment[did] = "truck_6"

    return assignment


def make_vehicle(vid: str, driver_name: str, vehicle_type: str = "truck_6") -> Vehicle:
    """Crea un Vehicle con los atributos derivados del vehicle_type."""
    cap_pallets = config.VEHICLE_CAPACITY_PALLETS.get(vehicle_type, 6)
    side_access = config.VEHICLE_HAS_SIDE_ACCESS.get(vehicle_type, True)
    lifo_mult = config.WH_LIFO_MULTIPLIER.get(vehicle_type, 1.0)
    return Vehicle(
        id=vid,
        driver_name=driver_name,
        capacity_boxes=cap_pallets * 60.0,
        start_lat=DEPOT_LAT,
        start_lon=DEPOT_LON,
        vehicle_type=vehicle_type,
        capacity_pallets=cap_pallets,
        has_side_access=side_access,
        lifo_penalty_multiplier=lifo_mult,
    )


# ── Main orders / vehicles loader ──────────────────────────────────────────────

def load_data(
    target_date: str | None = None,
    max_orders: int = 0,
) -> Tuple[List[Order], List[Vehicle]]:
    """
    Parse Hackaton.xlsx and return (orders, vehicles).

    - orders: one Order per unique *entrega* (aggregated across materials).
    - vehicles: one Vehicle per unique *repartidor*.
    - target_date: filter to a specific date string like '30/01/2026'.
    - max_orders: if >0, limit number of orders (useful for quick tests).
    """
    hackaton_path = config.HACKATON_XLSX
    horarios_path = config.HORARIOS_XLSX

    # ── Aux data ───────────────────────────────────────────────────────────────
    time_windows = load_time_windows(horarios_path)
    try:
        density = load_pallet_density()
    except Exception:
        density = {}

    from warehouse.classifier import classify

    # ── Open workbook ──────────────────────────────────────────────────────────
    wb = openpyxl.load_workbook(hackaton_path, read_only=True, data_only=True)
    ws_detail = wb["Detalle entrega"]

    # ── Aggregate lines by entrega id ─────────────────────────────────────────
    # Columns (0-indexed):
    # 0 FECHA | 1 Transporte | 2 Ruta | 3 Repartidor | 4 Destinatario mcía (name)
    # 5 Entrega | 6 Material | 7 Denom | 8 Cantidad | 9 UdMedida
    # 10 Destinatario mcía (id) | 11 Nombre1 | 12 Nombre2
    # 13 Calle | 14 CP | 15 Población | 16 ZonaTransp | 17 ZonaTransp

    orders_raw: Dict[str, dict] = {}   # entrega_id → aggregated dict
    drivers: Dict[str, str] = {}       # repartidor_id → name

    for row in ws_detail.iter_rows(min_row=2, values_only=True):
        fecha = str(row[0]) if row[0] else ""
        if target_date and fecha != target_date:
            continue

        entrega_id = str(row[5]) if row[5] else None
        driver_id = str(row[3]) if row[3] else None
        driver_name = str(row[4]) if row[4] else driver_id
        client_id = str(row[10]) if row[10] else ""
        material = str(row[6]) if row[6] else ""
        denom = str(row[7]) if row[7] else ""
        qty = _safe_float(row[8])

        if not entrega_id:
            continue

        if driver_id:
            drivers[driver_id] = driver_name

        if entrega_id not in orders_raw:
            orders_raw[entrega_id] = {
                "id": entrega_id,
                "driver_id": driver_id,
                "client_id": client_id,
                "client_name": str(row[11]) if row[11] else "",
                "address": str(row[13]) if row[13] else "",
                "postal_code": str(row[14]) if row[14] else "",
                "city": str(row[15]) if row[15] else "",
                "zone": str(row[16]) if row[16] else "",
                "total_boxes": 0.0,
                # Warehouse-aware
                "materials": set(),
                "categories": set(),
                "fragility_score": 0,
                "weight_score": 0,
                "has_returnables": False,
                "all_stackable": True,
                "pallet_fraction": 0.0,
            }
        d = orders_raw[entrega_id]
        d["total_boxes"] += qty
        if material:
            d["materials"].add(material)
            cls = classify(material, denom)
            d["categories"].add(cls["category"])
            if cls["fragility_score"] > d["fragility_score"]:
                d["fragility_score"] = cls["fragility_score"]
            if cls["weight_score"] > d["weight_score"]:
                d["weight_score"] = cls["weight_score"]
            if cls["is_returnable"]:
                d["has_returnables"] = True
            if not cls["stackable"]:
                d["all_stackable"] = False
            # Fracción de palet
            if material in density and density[material] > 0:
                d["pallet_fraction"] += qty / density[material]
            else:
                # Fallback: 1/80 por caja (densidad media)
                d["pallet_fraction"] += qty / 80.0

    wb.close()

    # ── Determine day of week from target_date ─────────────────────────────────
    dow = None
    if target_date:
        try:
            dt = datetime.strptime(target_date, "%d/%m/%Y")
            dow = dt.isoweekday()  # 1=Mon … 7=Sun
        except ValueError:
            pass

    # ── Build Order objects ────────────────────────────────────────────────────
    threshold = config.WH_PALLET_THRESHOLD
    orders: List[Order] = []
    for raw in list(orders_raw.values()):
        cid = raw["client_id"]
        tw_open, tw_close = 0, 86399
        if dow and cid in time_windows and dow in time_windows[cid]:
            tw_open, tw_close = time_windows[cid][dow]

        # Priority derived from time-window tightness (from Horarios Entrega.XLSX)
        # Tight window → high priority → bigger TW violation penalty in cost function
        tw_width = tw_close - tw_open  # seconds
        if tw_width < 3600:        # < 1 hour  (e.g. 10:30-11:00)
            priority = 1
        elif tw_width < 14400:     # 1 – 4 hours
            priority = 2
        else:                      # > 4 hours or all day (no real constraint)
            priority = 3

        frac = raw["pallet_fraction"]
        # Unload time proportional to order size: base + 7 min per full pallet
        service_time = max(
            config.UNLOAD_BASE_SECONDS,
            config.UNLOAD_BASE_SECONDS + int(frac * config.UNLOAD_PER_PALLET_S),
        )
        o = Order(
            id=raw["id"],
            client_id=cid,
            client_name=raw["client_name"],
            address=raw["address"],
            postal_code=raw["postal_code"],
            city=raw["city"],
            zone=raw["zone"],
            volume_boxes=raw["total_boxes"],
            tw_open=tw_open,
            tw_close=tw_close,
            service_time=service_time,
            priority=priority,
            n_references=len(raw["materials"]),
            categories=tuple(sorted(raw["categories"])),
            fragility_score=raw["fragility_score"] or 1,
            weight_score=raw["weight_score"] or 1,
            has_returnables=raw["has_returnables"],
            all_stackable=raw["all_stackable"],
            est_pallet_fraction=frac,
            needs_pallet=(frac >= threshold),
        )
        orders.append(o)
        if max_orders and len(orders) >= max_orders:
            break

    # ── Build Vehicle objects con tipos ────────────────────────────────────────
    vehicle_types = assign_vehicle_types(list(drivers.keys()))
    vehicles: List[Vehicle] = [
        make_vehicle(vid, name, vehicle_types.get(vid, "truck_6"))
        for vid, name in drivers.items()
    ]

    return orders, vehicles


# ── Geocoding helper (fallback: deterministic offset from depot) ───────────────

def geocode_orders(orders: List[Order], api_key: str = "") -> List[Order]:
    """
    Geocode orders using Google Maps Geocoding API.
    Falls back to a pseudo-coordinate based on postal-code hash when no key.
    """
    import requests

    for order in orders:
        if order.lat != 0.0 and order.lon != 0.0:
            continue

        full_address = f"{order.address}, {order.postal_code} {order.city}, Spain"

        if api_key:
            try:
                resp = requests.get(
                    "https://maps.googleapis.com/maps/api/geocode/json",
                    params={"address": full_address, "key": api_key},
                    timeout=5,
                )
                data = resp.json()
                if data.get("status") == "OK":
                    loc = data["results"][0]["geometry"]["location"]
                    order.lat = loc["lat"]
                    order.lon = loc["lng"]
                    continue
            except Exception:
                pass

        # Fallback: deterministic pseudo-coords from CP hash
        cp_hash = hash(order.postal_code) % 10000
        order.lat = DEPOT_LAT + (cp_hash % 100 - 50) * 0.001
        order.lon = DEPOT_LON + (cp_hash // 100 - 50) * 0.001

    return orders


def haversine_seconds(lat1, lon1, lat2, lon2, speed_kmh=35) -> float:
    """Travel time in seconds using haversine distance and average speed."""
    R = 6371.0
    phi1, phi2 = math.radians(lat1), math.radians(lat2)
    dphi = math.radians(lat2 - lat1)
    dlambda = math.radians(lon2 - lon1)
    a = math.sin(dphi / 2) ** 2 + math.cos(phi1) * math.cos(phi2) * math.sin(dlambda / 2) ** 2
    dist_km = 2 * R * math.asin(math.sqrt(a))
    return (dist_km / speed_kmh) * 3600

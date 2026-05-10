"""
run_week.py — Resuelve el VRP para los 5 días cacheados.
Sin llamadas a API. Usa time_matrix + fuel_matrix de data/.
Genera results_week.json para el frontend.
"""
from __future__ import annotations
import json, os, sys, copy
sys.path.insert(0, os.path.dirname(__file__))

import config

# SA rápido pero de calidad
config.SA_INITIAL_TEMP    = 3000.0
config.SA_COOLING_RATE    = 0.993
config.SA_MIN_TEMP        = 5.0
config.SA_MAX_ITER_PER_TEMP = 60

import openpyxl
from vrp.models   import Order, Vehicle, Trip, Solution
from vrp.savings  import clarke_wright
from vrp.annealing import simulated_annealing
from vrp.cost     import evaluate, BIG_M
import config as cfg

DAYS = ["16/03/2026","17/03/2026","18/03/2026","19/03/2026","20/03/2026"]
DEPOT_LAT, DEPOT_LON = cfg.DEPOT["lat"], cfg.DEPOT["lon"]

# Coste de combustible: 1 litro diesel ≈ 1.55 € → equivalencia en segundos
# Coste conductor ≈ 18 €/h → 1€ = 200s → 1 litro = 310s
FUEL_PENALTY_PER_LITRE = 310.0   # segundos equivalentes por litro


# ── Función de coste extendida (tiempo + combustible) ─────────────────────────

def trip_cost_combined(trip, orders, time_matrix, fuel_matrix, vehicle, order_idx):
    if not trip.stops:
        return 0.0, 0.0

    time_cost = 0.0
    fuel_total = 0.0
    prev = 0
    current_time = 0.0

    for stop in trip.stops:
        idx = order_idx.get(stop.id, 0)
        travel_t = time_matrix[prev][idx]
        fuel_l   = fuel_matrix[prev][idx]
        time_cost   += travel_t
        fuel_total  += fuel_l
        current_time += travel_t

        if current_time < stop.tw_open:
            current_time += stop.tw_open - current_time
        elif current_time > stop.tw_close:
            time_cost += (current_time - stop.tw_close) * stop.priority

        current_time += stop.service_time
        time_cost    += stop.service_time
        prev = idx

    # Vuelta al depósito
    time_cost  += time_matrix[prev][0]
    fuel_total += fuel_matrix[prev][0]
    current_time += time_matrix[prev][0]

    # Penalizaciones
    if trip.is_reload:
        time_cost += cfg.DEPOT_RELOAD_PENALTY_SECONDS

    load = sum(s.volume_boxes for s in trip.stops)
    if load > vehicle.capacity_boxes:
        time_cost += (load - vehicle.capacity_boxes) * 500

    if current_time > vehicle.max_route_seconds:
        time_cost += (current_time - vehicle.max_route_seconds) * 2

    trip.total_load_boxes = load
    combined = time_cost + fuel_total * FUEL_PENALTY_PER_LITRE
    return combined, fuel_total


def evaluate_combined(solution, orders, vehicles, time_matrix, fuel_matrix):
    vehicle_map = {v.id: v for v in vehicles}
    order_idx   = {o.id: i + 1 for i, o in enumerate(orders)}
    total_cost  = 0.0
    total_fuel  = 0.0
    reload_pen  = 0.0

    for trip in solution.trips:
        v = vehicle_map.get(trip.vehicle_id)
        if not v:
            continue
        c, f = trip_cost_combined(trip, orders, time_matrix, fuel_matrix, v, order_idx)
        total_cost += c
        total_fuel += f
        if trip.is_reload:
            reload_pen += cfg.DEPOT_RELOAD_PENALTY_SECONDS

    total_cost += len(solution.unserved_orders) * BIG_M
    solution.cost = total_cost
    solution.total_time_seconds = total_cost
    solution.total_reload_penalty = reload_pen
    solution._total_fuel = total_fuel
    return solution


# ── Loaders ────────────────────────────────────────────────────────────────────

def load_matrix(date_str):
    ddmmyyyy = date_str.replace("/","")
    path = f"data/matrix_{ddmmyyyy}.json"
    with open(path) as f:
        return json.load(f)


def load_orders_for_day(date_str, geocodes, matrix_data):
    """Construye Order objects usando geocodes + horarios."""
    from data.loader import load_time_windows
    tw = load_time_windows(cfg.HORARIOS_XLSX)

    from datetime import datetime
    dt = datetime.strptime(date_str, "%d/%m/%Y")
    dow = dt.isoweekday()  # 1=Mon…7=Sun

    client_ids = matrix_data["client_ids"]
    wb = openpyxl.load_workbook(cfg.HACKATON_XLSX, read_only=True, data_only=True)
    ws = wb["Detalle entrega"]

    # Aggregate boxes per entrega
    from collections import defaultdict
    raw = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if str(row[0]) != date_str:
            continue
        eid = str(row[5]) if row[5] else None
        cid = str(row[10]) if row[10] else None
        if not eid or cid not in client_ids:
            continue
        qty = float(row[8]) if row[8] else 0.0
        if eid not in raw:
            raw[eid] = {"client_id": cid, "name": str(row[11]) if row[11] else "",
                        "addr": str(row[13]) if row[13] else "",
                        "cp": str(row[14]) if row[14] else "",
                        "city": str(row[15]) if row[15] else "",
                        "zone": str(row[16]) if row[16] else "", "boxes": 0.0}
        raw[eid]["boxes"] += qty
    wb.close()

    # Agrupar por client_id — un Order por cliente, id = client_id
    # (la matriz está indexada por cliente, no por entrega)
    by_client = {}
    for eid, r in raw.items():
        cid = r["client_id"]
        if cid not in by_client:
            by_client[cid] = r.copy()
            by_client[cid]["boxes"] = 0.0
        by_client[cid]["boxes"] += r["boxes"]

    orders = []
    for cid, r in by_client.items():
        if cid not in client_ids:
            continue          # cliente sin posición en la matriz (no debería ocurrir)
        geo = geocodes.get(cid, {})
        tw_open, tw_close = 0, 86399
        if cid in tw and dow in tw[cid]:
            tw_open, tw_close = tw[cid][dow]
        orders.append(Order(
            id=cid, client_id=cid, client_name=r["name"],
            address=r["addr"], postal_code=r["cp"], city=r["city"], zone=r["zone"],
            lat=geo.get("lat", DEPOT_LAT), lon=geo.get("lon", DEPOT_LON),
            volume_boxes=r["boxes"], tw_open=tw_open, tw_close=tw_close,
        ))
    # Mantener el orden igual que en client_ids para que los índices coincidan con la matriz
    id_to_order = {o.id: o for o in orders}
    orders = [id_to_order[cid] for cid in client_ids if cid in id_to_order]
    return orders


def load_vehicles_for_day(date_str):
    wb = openpyxl.load_workbook(cfg.HACKATON_XLSX, read_only=True, data_only=True)
    ws = wb["Detalle entrega"]
    drivers = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if str(row[0]) == date_str and row[3]:
            drivers[str(row[3])] = str(row[4]) if row[4] else str(row[3])
    wb.close()
    return [Vehicle(id=vid, driver_name=name,
                    capacity_boxes=200, start_lat=DEPOT_LAT, start_lon=DEPOT_LON)
            for vid, name in drivers.items()]


# ── SA wrapper usando coste combinado ─────────────────────────────────────────

def solve_day(orders, vehicles, time_matrix, fuel_matrix):
    # Clarke-Wright solo necesita time_matrix (como lista de listas)
    initial = clarke_wright(orders, vehicles, time_matrix)
    evaluate_combined(initial, orders, vehicles, time_matrix, fuel_matrix)

    best = copy.deepcopy(initial)

    import math, random
    temp = config.SA_INITIAL_TEMP
    current = copy.deepcopy(initial)

    from vrp.annealing import _two_opt_move, _relocate_move, _swap_move

    while temp > config.SA_MIN_TEMP:
        for _ in range(config.SA_MAX_ITER_PER_TEMP):
            candidate = copy.deepcopy(current)
            trips = candidate.trips
            if not trips:
                break
            move = random.choice(["2opt","relocate","swap"])
            if move == "2opt":
                t = next((x for x in trips if len(x.stops)>1), None)
                if not t: continue
                nt = _two_opt_move(t, orders, time_matrix)
                if nt:
                    trips[trips.index(t)] = nt
            elif move == "relocate" and len(trips) >= 2:
                src, tgt = random.sample(trips, 2)
                res = _relocate_move(src, tgt, orders, time_matrix)
                if res:
                    new_src, new_tgt = res
                    for i,t in enumerate(trips):
                        if t is src: trips[i] = new_src
                        elif t is tgt: trips[i] = new_tgt
            elif move == "swap" and len(trips) >= 2:
                t1, t2 = random.sample(trips, 2)
                res = _swap_move(t1, t2)
                if res:
                    new_t1, new_t2 = res
                    for i,t in enumerate(trips):
                        if t is t1: trips[i] = new_t1
                        elif t is t2: trips[i] = new_t2

            candidate.trips = [t for t in trips if t.stops]
            evaluate_combined(candidate, orders, vehicles, time_matrix, fuel_matrix)

            delta = candidate.cost - current.cost
            if delta < 0 or random.random() < math.exp(-delta / temp):
                current = candidate
                if current.cost < best.cost:
                    best = copy.deepcopy(current)

        temp *= config.SA_COOLING_RATE

    return best


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    with open("data/geocodes.json") as f:
        geocodes = json.load(f)

    all_results = []

    for date_str in DAYS:
        print(f"\n{'='*60}")
        print(f"  {date_str}")
        print(f"{'='*60}")

        mdata      = load_matrix(date_str)
        time_mat   = mdata["time_matrix"]
        fuel_mat   = mdata["fuel_matrix"]
        orders     = load_orders_for_day(date_str, geocodes, mdata)
        vehicles   = load_vehicles_for_day(date_str)

        print(f"  {len(orders)} pedidos · {len(vehicles)} vehículos")

        solution = solve_day(orders, vehicles, time_mat, fuel_mat)

        # ── Métricas del día ──────────────────────────────────────────────────
        served      = sum(len(t.stops) for t in solution.trips)
        total_fuel  = getattr(solution, "_total_fuel", 0.0)
        total_secs  = sum(t.total_time_seconds for t in solution.trips)
        reloads     = solution.num_reloads()
        vehicles_used = solution.num_vehicles_used()

        h = int(total_secs)//3600
        m = (int(total_secs)%3600)//60
        print(f"  Vehículos usados: {vehicles_used}  |  Viajes: {solution.num_trips()}  |  Recargas: {reloads}")
        print(f"  Pedidos servidos: {served}/{len(orders)}  |  Tiempo total: {h}h {m:02d}m  |  Combustible: {total_fuel:.1f} L")

        # ── Serializar para el frontend ───────────────────────────────────────
        order_map = {o.id: o for o in orders}
        v_map     = {v.id: v for v in vehicles}

        from collections import defaultdict
        by_vehicle = defaultdict(list)
        for trip in solution.trips:
            by_vehicle[trip.vehicle_id].append(trip)

        routes_out = []
        for vid, trips in by_vehicle.items():
            v = v_map.get(vid)
            trip_list = []
            for i, trip in enumerate(trips):
                # Calcular fuel real de este viaje
                order_idx = {o.id: k+1 for k, o in enumerate(orders)}
                fuel_trip = 0.0
                prev = 0
                for stop in trip.stops:
                    idx = order_idx.get(stop.id, 0)
                    fuel_trip += fuel_mat[prev][idx]
                    prev = idx
                fuel_trip += fuel_mat[prev][0]

                trip_list.append({
                    "trip_number": i+1,
                    "is_reload": trip.is_reload,
                    "total_time_seconds": round(trip.total_time_seconds, 1),
                    "total_load_boxes": round(trip.total_load_boxes, 1),
                    "total_fuel_litres": round(fuel_trip, 2),
                    "stops": [
                        {
                            "order_id":    s.id,
                            "client_id":   s.client_id,
                            "client_name": s.client_name,
                            "city":        s.city,
                            "zone":        s.zone,
                            "boxes":       s.volume_boxes,
                            "lat":         order_map[s.id].lat,
                            "lon":         order_map[s.id].lon,
                            "tw_open":     s.tw_open,
                            "tw_close":    s.tw_close,
                        }
                        for s in trip.stops
                    ]
                })

            routes_out.append({
                "vehicle_id":   vid,
                "driver_name":  v.driver_name if v else vid,
                "trips":        trip_list,
                "total_fuel_litres": round(sum(t["total_fuel_litres"] for t in trip_list), 2),
                "total_time_seconds": round(sum(t["total_time_seconds"] for t in trip_list), 1),
            })

        all_results.append({
            "date":         date_str,
            "weekday":      ["Lunes","Martes","Miércoles","Jueves","Viernes"][DAYS.index(date_str)],
            "n_orders":     len(orders),
            "n_vehicles":   vehicles_used,
            "n_trips":      solution.num_trips(),
            "n_reloads":    reloads,
            "served":       served,
            "unserved":     len(solution.unserved_orders),
            "total_time_s": round(total_secs, 1),
            "total_fuel_l": round(total_fuel, 1),
            "depot":        cfg.DEPOT,
            "routes":       routes_out,
        })

    with open("results_week.json", "w", encoding="utf-8") as f:
        json.dump(all_results, f, indent=2, ensure_ascii=False)
    print("\n✓ results_week.json generado")


if __name__ == "__main__":
    main()

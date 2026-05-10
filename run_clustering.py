"""
run_clustering.py — Solves VRP for 5 cached days using logistics clustering.

Pipeline:
  1. Load cached time+fuel matrices (no API calls)
  2. Load and enrich orders (fragility, pallet estimates, categories)
  3. Build logistics distance matrix D[i,j] = α×T + β×F + γ×dist + δ×penalty
  4. Constrained K-Medoids clustering (one cluster per vehicle)
  5. Post-clustering refinement (move/swap heuristics)
  6. Assign warehouse zones (A–H by departure time)
  7. Route each cluster with greedy TSP + 2-opt + warehouse coupling
  8. Export results_clustering.json (same format as results_week.json)
"""
from __future__ import annotations
import argparse
import json
import os
import sys
import copy
from collections import defaultdict
from datetime import datetime
from typing import Dict, List, Tuple

sys.path.insert(0, os.path.dirname(__file__))

import openpyxl
import numpy as np

import config as cfg
from vrp.models import Order, Vehicle, Trip, Solution
from vrp.savings import estimate_pallets_byref
from vrp.cost import trip_cost
from data.loader import (
    load_time_windows, load_pallet_density, assign_vehicle_types, make_vehicle,
)
from warehouse.classifier import classify
from clustering.distance import build_logistics_matrix, build_depot_distances
from clustering.algorithm import cluster_orders
from clustering.heuristics import refine_clusters
from clustering.warehouse_zones import assign_warehouse_zones, warehouse_zone_report
from clustering.tsp import nearest_neighbor_tsp, two_opt_improve
from clustering.cost import cluster_fuel_litres, total_solution_cost
from clustering.models import Cluster, ClusterSolution


DAYS = ["16/03/2026", "17/03/2026", "18/03/2026", "19/03/2026", "20/03/2026"]
DEPOT_LAT, DEPOT_LON = cfg.DEPOT["lat"], cfg.DEPOT["lon"]
FUEL_TO_S = 310.0   # seconds equivalent per litre diesel


# ── Loaders ────────────────────────────────────────────────────────────────────

def load_matrix(date_str: str) -> dict:
    ddmmyyyy = date_str.replace("/", "")
    path = f"data/matrix_{ddmmyyyy}.json"
    with open(path) as f:
        return json.load(f)


def load_enriched_orders_by_client(
    date_str: str,
    geocodes: dict,
    matrix_data: dict,
) -> List[Order]:
    """
    Load orders for one day, grouped by client_id, enriched with:
    - Warehouse attributes: fragility_score, weight_score, categories, pallet_fraction
    - Time windows from HORARIOS XLSX
    - Geocoordinates from geocodes.json
    """
    tw = load_time_windows(cfg.HORARIOS_XLSX)
    try:
        density = load_pallet_density()
    except Exception:
        density = {}

    dt = datetime.strptime(date_str, "%d/%m/%Y")
    dow = dt.isoweekday()

    client_ids = matrix_data["client_ids"]
    client_id_set = set(client_ids)

    wb = openpyxl.load_workbook(cfg.HACKATON_XLSX, read_only=True, data_only=True)
    ws = wb["Detalle entrega"]

    # Aggregate per client (across all entregas for that client on this day)
    raw: Dict[str, dict] = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        if str(row[0]) != date_str:
            continue
        cid = str(row[10]) if row[10] else None
        if not cid or cid not in client_id_set:
            continue

        material = str(row[6]) if row[6] else ""
        denom    = str(row[7]) if row[7] else ""
        qty      = float(row[8]) if row[8] else 0.0

        if cid not in raw:
            raw[cid] = {
                "client_id":   cid,
                "client_name": str(row[11]) if row[11] else "",
                "address":     str(row[13]) if row[13] else "",
                "postal_code": str(row[14]) if row[14] else "",
                "city":        str(row[15]) if row[15] else "",
                "zone":        str(row[16]) if row[16] else "",
                "driver_id":   str(row[3]) if row[3] else None,
                "driver_name": str(row[4]) if row[4] else "",
                "total_boxes": 0.0,
                "materials":   set(),
                "categories":  set(),
                "fragility_score": 0,
                "weight_score":    0,
                "has_returnables": False,
                "all_stackable":   True,
                "pallet_fraction": 0.0,
            }

        d = raw[cid]
        d["total_boxes"] += qty
        if material:
            d["materials"].add(material)
            cls = classify(material, denom)
            d["categories"].add(cls["category"])
            if cls["fragility_score"] > d["fragility_score"]:
                d["fragility_score"] = cls["fragility_score"]
            if cls["weight_score"] > d["weight_score"]:
                d["weight_score"] = cls["weight_score"]
            if cls["is_returnable"]:
                d["has_returnables"] = True
            if not cls["stackable"]:
                d["all_stackable"] = False
            # Pallet fraction
            if material in density and density[material] > 0:
                d["pallet_fraction"] += qty / density[material]
            else:
                d["pallet_fraction"] += qty / 80.0

    wb.close()

    # Build Order objects, enriched
    threshold = cfg.WH_PALLET_THRESHOLD
    orders: List[Order] = []

    for cid, r in raw.items():
        geo = geocodes.get(cid, {})
        tw_open, tw_close = 0, 86399
        if cid in tw and dow in tw[cid]:
            tw_open, tw_close = tw[cid][dow]

        tw_width = tw_close - tw_open
        if tw_width < 3600:
            priority = 1
        elif tw_width < 14400:
            priority = 2
        else:
            priority = 3

        frac = r["pallet_fraction"]
        service_time = max(
            cfg.UNLOAD_BASE_SECONDS,
            cfg.UNLOAD_BASE_SECONDS + int(frac * cfg.UNLOAD_PER_PALLET_S),
        )
        park_range = cfg.PARKING_MAX_SECONDS - cfg.PARKING_MIN_SECONDS
        parking_time = cfg.PARKING_MIN_SECONDS + (abs(hash(cid)) % (park_range + 1))

        orders.append(Order(
            id=cid,
            client_id=cid,
            client_name=r["client_name"],
            address=r["address"],
            postal_code=r["postal_code"],
            city=r["city"],
            zone=r["zone"],
            lat=geo.get("lat", DEPOT_LAT),
            lon=geo.get("lon", DEPOT_LON),
            volume_boxes=r["total_boxes"],
            tw_open=tw_open,
            tw_close=tw_close,
            priority=priority,
            service_time=service_time,
            parking_time=parking_time,
            n_references=len(r["materials"]),
            categories=tuple(sorted(r["categories"])),
            fragility_score=r["fragility_score"] or 1,
            weight_score=r["weight_score"] or 1,
            has_returnables=r["has_returnables"],
            all_stackable=r["all_stackable"],
            est_pallet_fraction=frac,
            needs_pallet=(frac >= threshold),
        ))

    # Keep same order as client_ids so matrix indices are stable
    id_to_order = {o.id: o for o in orders}
    return [id_to_order[cid] for cid in client_ids if cid in id_to_order]


def load_vehicles_for_day(date_str: str) -> List[Vehicle]:
    """Load drivers and assign vehicle types deterministically."""
    wb = openpyxl.load_workbook(cfg.HACKATON_XLSX, read_only=True, data_only=True)
    ws = wb["Detalle entrega"]
    drivers: Dict[str, str] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if str(row[0]) == date_str and row[3]:
            drivers[str(row[3])] = str(row[4]) if row[4] else str(row[3])
    wb.close()

    vehicle_types = assign_vehicle_types(list(drivers.keys()))
    return [
        make_vehicle(vid, name, vehicle_types.get(vid, "truck_6"))
        for vid, name in drivers.items()
    ]


# ── Per-cluster routing ────────────────────────────────────────────────────────

def route_cluster(
    cluster: Cluster,
    time_matrix: List[List[float]],
    fuel_matrix: List[List[float]],
    matrix_idx: Dict[str, int],
    is_reload: bool = False,
    depot_idx: int = 0,
) -> Tuple[Trip, float]:
    """
    Build and cost a Trip for this cluster:
    greedy TSP → 2-opt → trip_cost (includes warehouse coupling).
    Returns (trip, fuel_litres).
    """
    v = cluster.vehicle
    if v is None or not cluster.orders:
        return None, 0.0

    stops = nearest_neighbor_tsp(cluster.orders, time_matrix, matrix_idx, depot_idx)
    stops = two_opt_improve(stops, time_matrix, matrix_idx, depot_idx)

    trip = Trip(vehicle_id=v.id, stops=stops, is_reload=is_reload)
    trip_cost(trip, cluster.orders, time_matrix, v)  # populates warehouse cached fields

    # Compute total route time: travel + parking + unload
    route_time = 0.0
    prev = depot_idx
    for stop in stops:
        mi = matrix_idx.get(stop.client_id, -1)
        if mi >= 0:
            route_time += time_matrix[prev][mi]
            prev = mi
        route_time += stop.parking_time + stop.service_time
    if prev != depot_idx:
        route_time += time_matrix[prev][depot_idx]
    if is_reload:
        route_time += cfg.DEPOT_RELOAD_PENALTY_SECONDS
    trip.total_time_seconds = route_time

    fuel = cluster_fuel_litres(cluster, fuel_matrix, matrix_idx, stops, depot_idx)
    return trip, fuel


# ── Day solver ────────────────────────────────────────────────────────────────

def solve_day_clustering(
    orders: List[Order],
    vehicles: List[Vehicle],
    time_matrix: List[List[float]],
    fuel_matrix: List[List[float]],
    matrix_data: dict,
) -> Tuple[ClusterSolution, List[Trip], float]:
    """
    Full clustering pipeline for one day.
    Returns (cluster_solution, trips, total_fuel_litres).
    """
    client_ids = matrix_data["client_ids"]
    matrix_idx: Dict[str, int] = {cid: i + 1 for i, cid in enumerate(client_ids)}
    # depot is index 0
    order_pos: Dict[str, int] = {o.id: i for i, o in enumerate(orders)}

    # ── Step 1: Build logistics distance matrix ───────────────────────────────
    D = build_logistics_matrix(time_matrix, fuel_matrix, orders, matrix_idx)
    depot_d = build_depot_distances(time_matrix, fuel_matrix, orders, matrix_idx)

    # ── Step 2: Cluster ────────────────────────────────────────────────────────
    solution = cluster_orders(orders, vehicles, D, order_pos, depot_d)

    # ── Step 3: Refine ─────────────────────────────────────────────────────────
    solution = refine_clusters(solution, D, order_pos, vehicles)

    # ── Step 4: Warehouse zones ────────────────────────────────────────────────
    assign_warehouse_zones(solution.clusters)

    # ── Step 5: Route each cluster ─────────────────────────────────────────────
    trips: List[Trip] = []
    total_fuel = 0.0
    vehicle_trip_count: Dict[str, int] = {}

    for cluster in solution.clusters:
        if cluster.vehicle is None:
            continue
        vid = cluster.vehicle.id
        is_reload = vehicle_trip_count.get(vid, 0) > 0
        vehicle_trip_count[vid] = vehicle_trip_count.get(vid, 0) + 1

        trip, fuel = route_cluster(
            cluster, time_matrix, fuel_matrix, matrix_idx,
            is_reload=is_reload,
        )
        if trip is not None:
            trips.append(trip)
            total_fuel += fuel
            cluster._trip = trip
            cluster._fuel = fuel

    # Unassigned → log
    if solution.unassigned:
        print(f"  ⚠ {len(solution.unassigned)} pedidos no asignados")

    # ── Post-process: enforce 8h daily cap using real routed times ─────────────
    trips, extra_unassigned = _enforce_daily_cap(trips, vehicles, time_matrix, matrix_idx)
    if extra_unassigned:
        print(f"  ⚠ {len(extra_unassigned)} paradas sin reubicar por cap diario 8h")
        solution.unassigned.extend(extra_unassigned)

    return solution, trips, total_fuel


def _enforce_daily_cap(
    trips: List[Trip],
    vehicles: List[Vehicle],
    time_matrix: List[List[float]],
    matrix_idx: Dict[str, int],
) -> Tuple[List[Trip], List[Order]]:
    """
    Hard enforce max_daily_seconds per vehicle using real routed times.

    1. For each overloaded vehicle: drop the excess trips (shortest first,
       preserving the longest/most important routes).
    2. Dropped stops → try to append to the last trip of any vehicle that
       still has daily capacity. Stops that can't be placed → unserved.
    """
    from collections import defaultdict

    vehicle_map = {v.id: v for v in vehicles}
    max_daily = {v.id: v.max_daily_seconds for v in vehicles}

    by_vehicle: Dict[str, List[Trip]] = defaultdict(list)
    for t in trips:
        by_vehicle[t.vehicle_id].append(t)

    kept: List[Trip] = []
    overflow_stops: List[Order] = []

    # ── Pass 1: keep trips that fit within each vehicle's daily cap ───────────
    for vid, v_trips in by_vehicle.items():
        cap = max_daily.get(vid, cfg.DEFAULT_MAX_DAILY_SECONDS)
        # Keep the longest trips first (main deliveries), drop shortest excess
        sorted_trips = sorted(v_trips, key=lambda t: -t.total_time_seconds)
        accumulated = 0.0
        for trip in sorted_trips:
            if accumulated + trip.total_time_seconds <= cap:
                kept.append(trip)
                accumulated += trip.total_time_seconds
            else:
                overflow_stops.extend(trip.stops)

    if not overflow_stops:
        return kept, []

    # ── Pass 2: redistribute overflow stops to vehicles with remaining time ───
    remaining_time: Dict[str, float] = {v.id: max_daily.get(v.id, cfg.DEFAULT_MAX_DAILY_SECONDS) for v in vehicles}
    for t in kept:
        remaining_time[t.vehicle_id] = remaining_time.get(t.vehicle_id, 0.0) - t.total_time_seconds

    still_unserved: List[Order] = []
    depot_idx = 0

    for stop in overflow_stops:
        stop_idx = matrix_idx.get(stop.client_id, 0)
        best_trip = None
        best_remaining = 0.0

        for t in kept:
            if not t.stops:
                continue
            last_idx = matrix_idx.get(t.stops[-1].client_id, depot_idx)
            # Conservative delta: travel to stop + service + parking + travel back to depot
            travel_in  = time_matrix[last_idx][stop_idx] if stop_idx > 0 else 600
            travel_out = time_matrix[stop_idx][depot_idx] if stop_idx > 0 else 600
            travel_was = time_matrix[last_idx][depot_idx] if last_idx > 0 else 0
            delta = travel_in + stop.parking_time + stop.service_time + travel_out - travel_was

            rem = remaining_time.get(t.vehicle_id, 0.0)
            # Only redistribute if it strictly fits within remaining daily time
            if rem >= delta and rem > best_remaining:
                best_remaining = rem
                best_trip = t
                best_delta = delta

        if best_trip is not None:
            best_trip.stops.append(stop)
            best_trip.total_time_seconds += best_delta
            remaining_time[best_trip.vehicle_id] -= best_delta
        else:
            still_unserved.append(stop)

    if still_unserved:
        print(f"  ⚠ {len(still_unserved)} paradas no redistribuibles (sin capacidad temporal)")

    return kept, still_unserved


# ── Serialiser ────────────────────────────────────────────────────────────────

def serialise_day(
    date_str: str,
    weekday: str,
    orders: List[Order],
    vehicles: List[Vehicle],
    solution: ClusterSolution,
    trips: List[Trip],
    total_fuel: float,
    fuel_matrix: List[List[float]],
    matrix_data: dict,
) -> dict:
    """Build the JSON dict for one day (same schema as results_week.json)."""
    client_ids = matrix_data["client_ids"]
    matrix_idx: Dict[str, int] = {cid: i + 1 for i, cid in enumerate(client_ids)}
    order_map  = {o.id: o for o in orders}
    v_map      = {v.id: v for v in vehicles}

    # Build trip→cluster map using object identity (cluster._trip may have been filtered out)
    trip_to_cluster: Dict[int, Cluster] = {}
    for cluster in solution.clusters:
        t = getattr(cluster, "_trip", None)
        if t is not None:
            trip_to_cluster[id(t)] = cluster

    # Group ONLY the kept trips by vehicle (respects _enforce_daily_cap filtering)
    by_vehicle: Dict[str, List] = defaultdict(list)
    for trip in trips:
        cluster = trip_to_cluster.get(id(trip))
        if cluster is not None and cluster.vehicle:
            by_vehicle[cluster.vehicle.id].append((trip, cluster))

    routes_out = []
    for vid, trip_cluster_pairs in by_vehicle.items():
        v = v_map.get(vid)
        trip_list = []
        for i, (trip, cluster) in enumerate(trip_cluster_pairs):
            # Compute fuel for this specific trip
            fuel_trip = 0.0
            prev = 0
            for stop in trip.stops:
                mi = matrix_idx.get(stop.client_id, -1)
                if mi >= 0:
                    fuel_trip += fuel_matrix[prev][mi]
                    prev = mi
            if prev > 0:
                fuel_trip += fuel_matrix[prev][0]

            trip_list.append({
                "trip_number":         i + 1,
                "is_reload":           trip.is_reload,
                "total_time_seconds":  round(trip.total_time_seconds, 1),
                "total_load_boxes":    round(trip.total_load_boxes, 1),
                "total_fuel_litres":   round(fuel_trip, 2),
                # Clustering / warehouse fields
                "cluster_id":          cluster.id,
                "warehouse_zone":      cluster.warehouse_zone,
                "n_pallets":           cluster.n_pallets,
                "chosen_strategy":     trip.chosen_strategy,
                "warehouse_cost_s":    round(trip.warehouse_cost, 1),
                "unload_multiplier":   trip.unload_multiplier,
                "stack_violations":    trip.stack_violations,
                "fragile_at_back_van": trip.fragile_at_back_van,
                "has_fragile":         cluster.has_fragile,
                "has_priority":        cluster.has_priority,
                "categories":          list(cluster.categories),
                "stops": [
                    {
                        "order_id":    s.id,
                        "client_id":   s.client_id,
                        "client_name": s.client_name,
                        "city":        s.city,
                        "zone":        s.zone,
                        "boxes":       s.volume_boxes,
                        "lat":         order_map[s.id].lat if s.id in order_map else DEPOT_LAT,
                        "lon":         order_map[s.id].lon if s.id in order_map else DEPOT_LON,
                        "tw_open":     s.tw_open,
                        "tw_close":    s.tw_close,
                        "fragility":   s.fragility_score,
                        "weight":      s.weight_score,
                        "n_pallets":   round(s.est_pallet_fraction, 2),
                    }
                    for s in trip.stops
                ],
            })

        routes_out.append({
            "vehicle_id":          vid,
            "driver_name":         v.driver_name if v else vid,
            "vehicle_type":        v.vehicle_type if v else "truck_6",
            "capacity_pallets":    v.capacity_pallets if v else 6,
            "trips":               trip_list,
            "total_fuel_litres":   round(sum(t["total_fuel_litres"] for t in trip_list), 2),
            "total_time_seconds":  round(sum(t["total_time_seconds"] for t in trip_list), 1),
        })

    # Warehouse zone summary
    zone_report = warehouse_zone_report(solution.clusters)

    served = sum(len(t["stops"]) for r in routes_out for t in r["trips"])
    total_time_s = sum(t["total_time_seconds"] for r in routes_out for t in r["trips"])

    return {
        "date":             date_str,
        "weekday":          weekday,
        "algorithm":        "logistics_clustering_kmedoids",
        "n_orders":         len(orders),
        "n_vehicles":       len({r["vehicle_id"] for r in routes_out}),
        "n_trips":          sum(len(r["trips"]) for r in routes_out),
        "n_clusters":       solution.num_clusters(),
        "served":           served,
        "unserved":         len(solution.unassigned),
        "total_time_s":     round(total_time_s, 1),
        "total_fuel_l":     round(total_fuel, 1),
        "co2_kg":           round(total_fuel * cfg.CO2_KG_PER_LITRE, 1),
        "warehouse_zones":  zone_report,
        "depot":            cfg.DEPOT,
        "routes":           routes_out,
        "unassigned_orders": [
            {"id": o.id, "city": o.city, "boxes": o.volume_boxes}
            for o in solution.unassigned
        ],
    }


# ── Main ──────────────────────────────────────────────────────────────────────

def main(dates: List[str] = None, output: str = "results_clustering.json"):
    with open("data/geocodes.json") as f:
        geocodes = json.load(f)

    target_days = dates or DAYS
    weekday_names = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes"]

    all_results = []

    for date_str in target_days:
        print(f"\n{'='*60}")
        print(f"  {date_str}  [clustering]")
        print(f"{'='*60}")

        mdata    = load_matrix(date_str)
        time_mat = mdata["time_matrix"]
        fuel_mat = mdata["fuel_matrix"]
        orders   = load_enriched_orders_by_client(date_str, geocodes, mdata)
        vehicles = load_vehicles_for_day(date_str)

        print(f"  {len(orders)} pedidos · {len(vehicles)} vehículos")

        solution, trips, total_fuel = solve_day_clustering(
            orders, vehicles, time_mat, fuel_mat, mdata
        )

        # Metrics
        served    = sum(len(c.orders) for c in solution.clusters)
        total_s   = sum(t.total_time_seconds for t in trips)
        n_fragile = sum(1 for c in solution.clusters if c.has_fragile)
        zone_rep  = warehouse_zone_report(solution.clusters)

        h = int(total_s) // 3600
        m = (int(total_s) % 3600) // 60
        print(f"  Clusters: {solution.num_clusters()}  |  Servidos: {served}/{len(orders)}")
        print(f"  Tiempo total: {h}h {m:02d}m  |  Combustible: {total_fuel:.1f}L  "
              f"| CO2: {total_fuel * cfg.CO2_KG_PER_LITRE:.1f}kg")
        print(f"  Clusters con frágiles: {n_fragile}  |  Zonas: {list(zone_rep.keys())}")

        day_idx = DAYS.index(date_str) if date_str in DAYS else 0
        result = serialise_day(
            date_str,
            weekday_names[day_idx],
            orders,
            vehicles,
            solution,
            trips,
            total_fuel,
            fuel_mat,
            mdata,
        )
        all_results.append(result)

    with open(output, "w", encoding="utf-8") as f:
        json.dump(all_results, f, indent=2, ensure_ascii=False)
    print(f"\n✓ {output} generado")


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--date", help="Single date DD/MM/YYYY (default: all 5 days)")
    parser.add_argument("--output", default="results_clustering.json")
    args = parser.parse_args()

    dates = [args.date] if args.date else None
    main(dates=dates, output=args.output)
